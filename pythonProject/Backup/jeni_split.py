import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

# 1. Configuration
file_path = 'Masters.xlsx'  # Make sure this file is SAVED

# 2. Find the Header Row
print("Reading file...")
df_raw = pd.read_excel(file_path, header=None)

header_row_index = None
# We check the first 15 rows for any of these "anchor" words
anchors = ['talent_id', 'sl no', 'sno', 'talent id', 'sl. no.']

for i, row in df_raw.iterrows():
    row_str = row.astype(str).str.lower().values
    if any(anchor in "".join(row_str) for anchor in anchors):
        header_row_index = i
        break

if header_row_index is None:
    print("❌ Still can't find the header. Here is what Row 2 looks like:")
    print(df_raw.iloc[2].values)
else:
    # 3. Load data with the found header
    df = pd.read_excel(file_path, header=header_row_index)
    # Clean column names: remove spaces, dots, and make lowercase
    df.columns = df.columns.astype(str).str.strip().str.lower().str.replace('.', '', regex=False).str.replace(' ', '_',
                                                                                                              regex=False)

    print(f"✅ Found headers at row {header_row_index + 1}")
    print(f"I am using these columns: {df.columns.tolist()}")


    # 4. Map Columns Automatically
    # This looks for the best match so it doesn't crash if the name is slightly off
    def get_col(options):
        for opt in options:
            if opt in df.columns: return opt
        return None


    col_id = get_col(['talent_id', 'employee_id', 'sl_no'])
    col_name = get_col(['talent_name', 'name', 'employee_name'])
    col_date = get_col(['date', 'date_of_joining', 'doj'])
    col_team = get_col(['team', 'business_function', 'department', 'business_vertical'])

    # 5. Create the Cards
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Cards"

    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    card_index = 0
    for _, row in df.iterrows():
        if pd.isna(row.get(col_id)): continue  # Skip empty rows

        r_start = (card_index // 2) * 5 + 1
        c_start = (card_index % 2) * 3 + 1

        data_map = [
            ("Name :", row.get(col_name, "N/A")),
            ("Employee ID :", row.get(col_id, "N/A")),
            ("Joining Date :", row.get(col_date, "N/A")),
            ("Business Vertical / Enabler :", row.get(col_team, "N/A"))
        ]

        for i, (label, value) in enumerate(data_map):
            cl = ws.cell(row=r_start + i, column=c_start, value=label)
            cl.font, cl.border = bold_font, thin_border
            cv = ws.cell(row=r_start + i, column=c_start + 1, value=value)
            cv.border = thin_border
            if i == 2: cv.number_format = 'DD-MMM-YY'

        card_index += 1

    for col in ['A', 'B', 'D', 'E']: ws.column_dimensions[col].width = 28
    wb.save('Formatted_Employee_Cards.xlsx')
    print(f"🚀 Success! Created {card_index} cards.")